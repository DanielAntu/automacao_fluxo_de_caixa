from pathlib import Path
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

USERS = os.environ.get('USERS', '')
PASSWORD = os.environ.get('PASSWORD', '')

ROOT_PATCH = Path(__file__).parent
CHROMEDRIVER_NAME = 'chromedriver.exe'
CHROMEDRIVER_PATH = ROOT_PATCH / 'bin' / CHROMEDRIVER_NAME

def make_browser(*options):
    chrome_options = webdriver.ChromeOptions()

    if options is not None:
        for option in options:
            chrome_options.add_argument(option)

    chrome_service = Service(executable_path=CHROMEDRIVER_PATH)
    browser = webdriver.Chrome(service=chrome_service, options=chrome_options)

    return browser

def preacherForm(browser, xpath, valor):
    browser.find_element(By.XPATH, xpath).send_keys(valor)

# openpyxl
workbook = load_workbook('dados_registro.xlsx')
plan = workbook['Sheet1']

browser = make_browser()
browser.get('http://127.0.0.1:8000/')

# login
preacherForm(browser, '//*[@id="id_username"]', USERS)
sleep(1)
preacherForm(browser, '//*[@id="id_password"]', PASSWORD)
sleep(1)
browser.find_element(By.XPATH, '/html/body/main/div/form/input[2]').click()

# registrar fluxo
for linha in plan.iter_rows(min_row=2, values_only=True):
    descricao, natureza, tipo, valor = linha
    preacherForm(browser, '//*[@id="id_description"]', descricao)
    preacherForm(browser, '//*[@id="id_nature"]', natureza)
    if tipo is not None:
        preacherForm(browser, '//*[@id="id_type_cash"]', tipo)
    preacherForm(browser, '//*[@id="id_value"]', valor)
    browser.find_element(By.XPATH, '/html/body/main/div[2]/form/input[2]').click()
    sleep(2)
    

sleep(5)
browser.quit()